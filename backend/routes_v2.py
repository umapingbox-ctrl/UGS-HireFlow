"""UGS HireFlow — Phase 2 extension routes.
Adds: partners, comprehensive registration with files, verification workflow,
soft delete / archive / merge, employee workload, company hiring history,
job workspace actions, batch history, organization settings, saved filters,
Excel .xlsx import/export, notification triggers.
Imports 'db' from db module and reuses helpers.
"""
import io
import os
import logging
from datetime import datetime, timezone
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from db import db, log_activity, create_notification
from auth import get_current_user, require_roles, hash_password
from storage import put_object, APP_NAME
from models import (
    Candidate, CandidateCreate, Partner, PartnerBase,
    VerifyCandidateRequest, OrganizationSettings, SavedFilter,
    MergeCandidatesRequest, FileRef, PaymentRecord,
    now_iso, new_id,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["v2"])


# ================================================================
# HELPERS
# ================================================================
async def next_code(collection: str, prefix: str) -> str:
    count = await db[collection].count_documents({})
    return f"{prefix}-{count + 1:04d}"


def compute_profile_completion(c: dict) -> int:
    """Score 0-100 based on filled fields."""
    weights = {
        "full_name": 5, "email": 5, "phone": 5, "city": 3, "state": 2,
        "date_of_birth": 3, "gender": 2, "address": 3,
        "total_experience_years": 5, "current_salary": 3, "expected_salary": 3,
        "notice_period_days": 3, "photo_file_id": 8, "resume_file_id": 12,
    }
    list_weights = {"education": 10, "experience": 8, "skills": 8,
                     "preferred_locations": 3, "documents": 5}
    dict_weights = {"reference": 3, "emergency_contact": 2}
    score = 0
    for k, w in weights.items():
        if c.get(k): score += w
    for k, w in list_weights.items():
        if c.get(k) and len(c.get(k) or []) > 0: score += w
    for k, w in dict_weights.items():
        v = c.get(k)
        if v and (v.get("name") if isinstance(v, dict) else v): score += w
    return min(100, score)


async def upload_file_for_candidate(file: UploadFile, candidate_id: str,
                                      label: str, owner_id: str) -> dict:
    """Upload a file to object storage & DB, return FileRef dict."""
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
    path = f"{APP_NAME}/uploads/{candidate_id}/{new_id()}.{ext}"
    data = await file.read()
    result = put_object(path, data, file.content_type or "application/octet-stream")
    file_ref = FileRef(storage_path=result["path"],
                        original_filename=file.filename,
                        content_type=file.content_type or "application/octet-stream",
                        size=result.get("size", len(data)),
                        owner_id=owner_id)
    await db.files.insert_one(file_ref.model_dump())
    return file_ref.model_dump()


# ================================================================
# COMPREHENSIVE REGISTRATION (multipart with files)
# ================================================================
@router.post("/public/register-candidate-full")
async def public_register_full(
    full_name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    password: Optional[str] = Form(None),
    alternate_phone: Optional[str] = Form(None),
    date_of_birth: Optional[str] = Form(None),
    gender: Optional[str] = Form(None),
    address: Optional[str] = Form(None),
    city: Optional[str] = Form(None),
    state: Optional[str] = Form(None),
    country: Optional[str] = Form("India"),
    pincode: Optional[str] = Form(None),
    skills: Optional[str] = Form(""),  # comma-separated
    total_experience_years: Optional[float] = Form(0),
    current_salary: Optional[float] = Form(None),
    expected_salary: Optional[float] = Form(None),
    notice_period_days: Optional[int] = Form(30),
    preferred_locations: Optional[str] = Form(""),  # comma-separated
    education_degree: Optional[str] = Form(None),
    education_institution: Optional[str] = Form(None),
    education_year: Optional[str] = Form(None),
    experience_company: Optional[str] = Form(None),
    experience_role: Optional[str] = Form(None),
    experience_duration: Optional[str] = Form(None),
    reference_name: Optional[str] = Form(None),
    reference_phone: Optional[str] = Form(None),
    emergency_name: Optional[str] = Form(None),
    emergency_phone: Optional[str] = Form(None),
    photo: Optional[UploadFile] = File(None),
    resume: Optional[UploadFile] = File(None),
    certificates: List[UploadFile] = File(default_factory=list),
    experience_documents: List[UploadFile] = File(default_factory=list),
    supporting_documents: List[UploadFile] = File(default_factory=list),
):
    email = email.lower().strip()
    if await db.candidates.find_one({"email": email}):
        raise HTTPException(400, "Email already registered")

    # Build candidate
    cand = Candidate(
        full_name=full_name, email=email, phone=phone,
        alternate_phone=alternate_phone, date_of_birth=date_of_birth,
        gender=gender, address=address, city=city, state=state,
        country=country, pincode=pincode,
        skills=[s.strip() for s in (skills or "").split(",") if s.strip()],
        total_experience_years=total_experience_years or 0,
        current_salary=current_salary, expected_salary=expected_salary,
        notice_period_days=notice_period_days,
        preferred_locations=[s.strip() for s in (preferred_locations or "").split(",") if s.strip()],
        education=[{"degree": education_degree, "institution": education_institution,
                     "year": education_year}] if education_degree else [],
        experience=[{"company": experience_company, "role": experience_role,
                      "duration": experience_duration}] if experience_company else [],
        reference={"name": reference_name, "phone": reference_phone} if reference_name else None,
        emergency_contact={"name": emergency_name, "phone": emergency_phone} if emergency_name else None,
        reference_name=reference_name,
        reference_phone=reference_phone,
    )
    cand.candidate_code = await next_code("candidates", "UGS-C")

    # Upload files
    if photo:
        fr = await upload_file_for_candidate(photo, cand.id, "photo", cand.id)
        cand.photo_file_id = fr["id"]
    if resume:
        fr = await upload_file_for_candidate(resume, cand.id, "resume", cand.id)
        cand.resume_file_id = fr["id"]
    docs = []
    for lbl, files in [("certificate", certificates or []),
                         ("experience_document", experience_documents or []),
                         ("supporting_document", supporting_documents or [])]:
        for f in files:
            if f and f.filename:
                fr = await upload_file_for_candidate(f, cand.id, lbl, cand.id)
                docs.append({"id": new_id(), "label": lbl, "file_id": fr["id"],
                              "uploaded_at": now_iso()})
    cand.documents = docs
    cand.profile_completion = compute_profile_completion(cand.model_dump())

    await db.candidates.insert_one(cand.model_dump())

    # Optional login account
    if password:
        await db.users.insert_one({
            "id": new_id(), "email": cand.email, "full_name": cand.full_name,
            "role": "candidate", "phone": cand.phone, "is_active": True,
            "candidate_id": cand.id,
            "password_hash": hash_password(password),
            "created_at": now_iso(),
        })

    await log_activity(None, "candidate.registered",
                       f"New candidate registered: {cand.full_name}",
                       "candidate", cand.id)

    # Notify admins
    async for admin in db.users.find({"role": "admin"}):
        await create_notification(admin["id"], "New Candidate Registration",
                                    f"{cand.full_name} has registered and awaits verification.",
                                    "info", f"/app/candidates/{cand.id}")

    d = cand.model_dump()
    d.pop("_id", None)
    return d


# ================================================================
# PARTNERS
# ================================================================
@router.get("/partners")
async def list_partners(q: str = "", current: dict = Depends(get_current_user)):
    filt = {}
    if q:
        filt = {"$or": [{"name": {"$regex": q, "$options": "i"}},
                         {"phone": {"$regex": q, "$options": "i"}},
                         {"partner_code": {"$regex": q, "$options": "i"}}]}
    items = await db.partners.find(filt, {"_id": 0}).sort("created_at", -1).to_list(500)
    # attach candidate counts
    for p in items:
        p["candidate_count"] = await db.candidates.count_documents(
            {"partner_id": p["id"], "is_deleted": {"$ne": True}})
    return items


@router.get("/partners/{pid}")
async def get_partner(pid: str, current: dict = Depends(get_current_user)):
    p = await db.partners.find_one({"id": pid}, {"_id": 0})
    if not p: raise HTTPException(404)
    p["candidates"] = await db.candidates.find(
        {"partner_id": pid, "is_deleted": {"$ne": True}},
        {"_id": 0}).sort("created_at", -1).to_list(500)
    p["candidate_count"] = len(p["candidates"])
    return p


@router.post("/partners")
async def create_partner(body: PartnerBase,
                          current: dict = Depends(require_roles(["admin"]))):
    existing = await db.partners.find_one({"phone": body.phone})
    if existing:
        existing.pop("_id", None)
        return existing  # idempotent by phone
    p = Partner(**body.model_dump(), created_by=current["id"])
    p.partner_code = await next_code("partners", "UGS-P")
    await db.partners.insert_one(p.model_dump())
    await log_activity(current, "partner.created",
                       f"Created partner {p.name}", "partner", p.id)
    return p.model_dump()


@router.patch("/partners/{pid}")
async def update_partner(pid: str, body: dict,
                          current: dict = Depends(require_roles(["admin"]))):
    body.pop("id", None); body.pop("_id", None)
    await db.partners.update_one({"id": pid}, {"$set": body})
    await log_activity(current, "partner.updated",
                       f"Updated partner {pid}", "partner", pid)
    return await db.partners.find_one({"id": pid}, {"_id": 0})


@router.delete("/partners/{pid}")
async def delete_partner(pid: str, current: dict = Depends(require_roles(["admin"]))):
    count = await db.candidates.count_documents({"partner_id": pid, "is_deleted": {"$ne": True}})
    if count > 0:
        raise HTTPException(400, f"Cannot delete: partner has {count} linked candidates")
    await db.partners.delete_one({"id": pid})
    return {"success": True}


# ================================================================
# VERIFY CANDIDATE (with partner picker / creator)
# ================================================================
@router.post("/candidates/{cid}/verify-v2")
async def verify_candidate_v2(cid: str, body: VerifyCandidateRequest,
                                current: dict = Depends(require_roles(["admin"]))):
    c = await db.candidates.find_one({"id": cid})
    if not c: raise HTTPException(404)

    partner_id = body.partner_id
    if body.new_partner:
        # dedupe by phone
        existing = await db.partners.find_one({"phone": body.new_partner.phone})
        if existing:
            partner_id = existing["id"]
        else:
            p = Partner(**body.new_partner.model_dump(), created_by=current["id"])
            p.partner_code = await next_code("partners", "UGS-P")
            await db.partners.insert_one(p.model_dump())
            partner_id = p.id
            await log_activity(current, "partner.created",
                                f"Created partner {p.name} from verification",
                                "partner", p.id)

    updates = {"status": "verified", "verified_at": now_iso(), "updated_at": now_iso()}
    if partner_id:
        updates["partner_id"] = partner_id
    await db.candidates.update_one({"id": cid}, {"$set": updates})
    await log_activity(current, "candidate.verified",
                       f"Verified {c['full_name']}{' (partner linked)' if partner_id else ''}",
                       "candidate", cid, {"partner_id": partner_id, "remarks": body.remarks})

    # Notify recruiters / candidate
    cu = await db.users.find_one({"candidate_id": cid})
    if cu:
        await create_notification(cu["id"], "Profile Verified",
                                    "Your profile has been verified.",
                                    "success", "/me")
    return {"success": True, "partner_id": partner_id}


# ================================================================
# SOFT DELETE / ARCHIVE / RESTORE / PERMANENT DELETE
# ================================================================
@router.post("/candidates/{cid}/archive")
async def archive_candidate(cid: str, current: dict = Depends(require_roles(["admin", "employee"]))):
    r = await db.candidates.update_one({"id": cid, "is_deleted": {"$ne": True}},
        {"$set": {"is_archived": True, "archived_at": now_iso(), "updated_at": now_iso()}})
    if r.matched_count == 0: raise HTTPException(404)
    await log_activity(current, "candidate.archived", f"Archived candidate {cid}",
                        "candidate", cid)
    return {"success": True}


@router.post("/candidates/{cid}/restore")
async def restore_candidate(cid: str, current: dict = Depends(require_roles(["admin"]))):
    await db.candidates.update_one({"id": cid},
        {"$set": {"is_archived": False, "is_deleted": {"$ne": True},
                    "archived_at": None, "deleted_at": None, "updated_at": now_iso()}})
    await log_activity(current, "candidate.restored", f"Restored candidate {cid}",
                        "candidate", cid)
    return {"success": True}


@router.post("/candidates/{cid}/soft-delete")
async def soft_delete_candidate(cid: str, current: dict = Depends(require_roles(["admin"]))):
    await db.candidates.update_one({"id": cid},
        {"$set": {"is_deleted": True, "deleted_at": now_iso(), "updated_at": now_iso()}})
    await log_activity(current, "candidate.soft_deleted", f"Soft-deleted candidate {cid}",
                        "candidate", cid)
    return {"success": True}


@router.delete("/candidates/{cid}/permanent")
async def permanent_delete_candidate(cid: str, current: dict = Depends(require_roles(["admin"]))):
    c = await db.candidates.find_one({"id": cid})
    if not c: raise HTTPException(404)
    await db.candidates.delete_one({"id": cid})
    await log_activity(current, "candidate.deleted",
                        f"Permanently deleted candidate {c.get('full_name')}",
                        "candidate", cid)
    return {"success": True}


# ================================================================
# DUPLICATE DETECTION + MERGE
# ================================================================
@router.get("/candidates/duplicates/check")
async def check_duplicates(email: str = "", phone: str = "",
                             current: dict = Depends(get_current_user)):
    filt = {"is_deleted": {"$ne": True}, "$or": []}
    if email: filt["$or"].append({"email": email.lower()})
    if phone: filt["$or"].append({"phone": phone})
    if not filt["$or"]: return []
    items = await db.candidates.find(filt,
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "phone": 1,
          "candidate_code": 1, "status": 1}).to_list(20)
    return items


@router.post("/candidates/merge")
async def merge_candidates(body: MergeCandidatesRequest,
                             current: dict = Depends(require_roles(["admin"]))):
    src = await db.candidates.find_one({"id": body.source_id})
    tgt = await db.candidates.find_one({"id": body.target_id})
    if not src or not tgt: raise HTTPException(404)
    if src["id"] == tgt["id"]: raise HTTPException(400, "Cannot merge with self")

    # Merge payments, timeline, notes, documents, skills, preferred_locations
    push = {}
    for k in ["payments", "interview_timeline", "notes", "documents"]:
        if src.get(k):
            push.setdefault("$push", {})[k] = {"$each": src[k]}
    # Add to sets
    add = {}
    for k in ["skills", "preferred_locations"]:
        if src.get(k):
            add[k] = {"$each": src[k]}
    if add: push.setdefault("$addToSet", {}).update(add)
    # Aggregate amounts
    inc = {"amount_paid": src.get("amount_paid") or 0}
    if inc["amount_paid"]:
        push.setdefault("$inc", {}).update(inc)
    push.setdefault("$set", {})["updated_at"] = now_iso()

    if push:
        await db.candidates.update_one({"id": tgt["id"]}, push)

    # Soft-delete source
    await db.candidates.update_one({"id": src["id"]},
        {"$set": {"is_deleted": True, "deleted_at": now_iso(),
                    "notes": (src.get("notes") or []) + [
                        {"id": new_id(),
                          "text": f"Merged into {tgt['candidate_code']} ({tgt['full_name']})",
                          "author": current["email"], "created_at": now_iso()}
                    ]}})
    await log_activity(current, "candidate.merged",
                        f"Merged {src['full_name']} into {tgt['full_name']}",
                        "candidate", tgt["id"], {"source_id": src["id"]})
    return {"success": True}


# ================================================================
# PROFILE COMPLETION recompute
# ================================================================
@router.post("/candidates/{cid}/recompute-completion")
async def recompute_completion(cid: str, current: dict = Depends(get_current_user)):
    c = await db.candidates.find_one({"id": cid}, {"_id": 0})
    if not c: raise HTTPException(404)
    pct = compute_profile_completion(c)
    await db.candidates.update_one({"id": cid}, {"$set": {"profile_completion": pct}})
    return {"profile_completion": pct}


@router.get("/candidates/{cid}/quick")
async def quick_view(cid: str, current: dict = Depends(get_current_user)):
    c = await db.candidates.find_one({"id": cid},
        {"_id": 0, "full_name": 1, "candidate_code": 1, "email": 1,
          "phone": 1, "status": 1, "payment_status": 1, "skills": 1,
          "total_experience_years": 1, "expected_salary": 1,
          "assigned_employee_id": 1, "assigned_company_id": 1,
          "profile_completion": 1, "id": 1})
    if not c: raise HTTPException(404)
    return c


# ================================================================
# EMPLOYEE DETAIL — workload / performance / assigned candidates
# ================================================================
@router.get("/employees/{eid}")
async def employee_detail(eid: str, current: dict = Depends(require_roles(["admin", "employee"]))):
    if current["role"] == "employee" and current["id"] != eid:
        raise HTTPException(403)
    u = await db.users.find_one({"id": eid, "role": "employee"},
                                  {"_id": 0, "password_hash": 0})
    if not u: raise HTTPException(404)

    filt = {"assigned_employee_id": eid, "is_deleted": {"$ne": True}}
    assigned = await db.candidates.find(filt, {"_id": 0}).sort("created_at", -1).to_list(500)
    placed = [c for c in assigned if c.get("status") == "placed"]
    active_pipeline = [c for c in assigned if c.get("status") not in
                        ["placed", "rejected", "inactive"]]
    stage_counts = {}
    for c in assigned:
        stage_counts[c.get("status", "unknown")] = stage_counts.get(c.get("status"), 0) + 1

    logs = await db.activity_logs.find({"actor_id": eid}, {"_id": 0}) \
        .sort("created_at", -1).limit(50).to_list(50)

    login_history = await db.activity_logs.find(
        {"actor_id": eid, "action": {"$in": ["user.login", "user.logout"]}},
        {"_id": 0}).sort("created_at", -1).limit(20).to_list(20)

    return {
        **u,
        "workload": {"assigned": len(assigned), "placed": len(placed),
                       "active_pipeline": len(active_pipeline)},
        "stage_counts": stage_counts,
        "assigned_candidates": assigned,
        "activity": logs,
        "login_history": login_history,
    }


# ================================================================
# COMPANY DETAIL EXTENDED — hiring history
# ================================================================
@router.get("/companies/{cid}/hiring-history")
async def company_hiring_history(cid: str, current: dict = Depends(get_current_user)):
    co = await db.companies.find_one({"id": cid}, {"_id": 0})
    if not co: raise HTTPException(404)
    jobs = await db.jobs.find({"company_id": cid}, {"_id": 0}) \
        .sort("created_at", -1).to_list(500)
    total_placed = await db.candidates.count_documents({
        "assigned_company_id": cid, "status": "placed", "is_deleted": {"$ne": True}})
    active_candidates = await db.candidates.count_documents({
        "assigned_company_id": cid,
        "status": {"$nin": ["placed", "rejected", "inactive"]},
        "is_deleted": {"$ne": True}})
    open_jobs = sum(1 for j in jobs if j.get("status") == "open")
    return {
        "company": co,
        "stats": {"total_jobs": len(jobs), "open_jobs": open_jobs,
                    "total_placed": total_placed,
                    "active_candidates": active_candidates},
        "jobs": jobs,
    }


# ================================================================
# JOB ACTIONS — close / reopen / bulk allocate / remove candidate
# ================================================================
@router.post("/jobs/{jid}/close")
async def close_job(jid: str, current: dict = Depends(require_roles(["admin", "employee"]))):
    await db.jobs.update_one({"id": jid},
        {"$set": {"status": "closed", "closed_at": now_iso()}})
    await log_activity(current, "job.closed", f"Closed job {jid}", "job", jid)
    return {"success": True}


@router.post("/jobs/{jid}/reopen")
async def reopen_job(jid: str, current: dict = Depends(require_roles(["admin", "employee"]))):
    await db.jobs.update_one({"id": jid},
        {"$set": {"status": "open", "reopened_at": now_iso()}})
    await log_activity(current, "job.reopened", f"Reopened job {jid}", "job", jid)
    return {"success": True}


@router.post("/jobs/{jid}/remove-candidate")
async def remove_candidate_from_job(jid: str, candidate_id: str = Form(...),
                                      current: dict = Depends(require_roles(["admin", "employee"]))):
    await db.jobs.update_one({"id": jid},
        {"$pull": {"assigned_candidate_ids": candidate_id}})
    await db.candidates.update_one({"id": candidate_id, "assigned_job_id": jid},
        {"$set": {"assigned_job_id": None, "updated_at": now_iso()}})
    await log_activity(current, "job.candidate_removed",
                        f"Removed candidate {candidate_id} from job {jid}",
                        "job", jid)
    return {"success": True}


@router.get("/jobs/{jid}/workspace")
async def job_workspace(jid: str, current: dict = Depends(get_current_user)):
    j = await db.jobs.find_one({"id": jid}, {"_id": 0})
    if not j: raise HTTPException(404)
    co = await db.companies.find_one({"id": j["company_id"]}, {"_id": 0})
    candidates = await db.candidates.find(
        {"id": {"$in": j.get("assigned_candidate_ids", [])}, "is_deleted": {"$ne": True}},
        {"_id": 0}).to_list(500)
    stage_counts = {}
    for c in candidates:
        stage_counts[c.get("status", "unknown")] = stage_counts.get(c.get("status"), 0) + 1
    placed = sum(1 for c in candidates if c.get("status") == "placed")
    return {"job": j, "company": co, "candidates": candidates,
             "stats": {"allocated": len(candidates), "placed": placed,
                         "stage_counts": stage_counts,
                         "vacancies_remaining": max(0, (j.get("vacancies", 1) or 0) - placed)}}


# ================================================================
# BATCH ACTIONS — status transitions
# ================================================================
@router.post("/batches/{bid}/mark-running")
async def mark_batch_running(bid: str, current: dict = Depends(require_roles(["admin"]))):
    await db.batches.update_one({"id": bid}, {"$set": {"status": "running"}})
    await log_activity(current, "batch.status_changed", f"Batch {bid} → running",
                        "batch", bid)
    return {"success": True}


@router.post("/batches/{bid}/mark-completed")
async def mark_batch_completed(bid: str, current: dict = Depends(require_roles(["admin"]))):
    await db.batches.update_one({"id": bid},
        {"$set": {"status": "completed", "completed_at": now_iso()}})
    await log_activity(current, "batch.status_changed", f"Batch {bid} → completed",
                        "batch", bid)
    return {"success": True}


@router.post("/batches/{bid}/remove-candidate")
async def remove_candidate_from_batch(bid: str, candidate_id: str = Form(...),
                                        current: dict = Depends(require_roles(["admin"]))):
    await db.batches.update_one({"id": bid}, {"$pull": {"candidate_ids": candidate_id}})
    await db.candidates.update_one({"id": candidate_id},
        {"$set": {"batch_id": None, "updated_at": now_iso()}})
    return {"success": True}


# ================================================================
# ORGANIZATION SETTINGS
# ================================================================
@router.get("/settings/organization")
async def get_org_settings(current: dict = Depends(get_current_user)):
    s = await db.settings.find_one({"id": "org_settings"}, {"_id": 0})
    if not s:
        default = OrganizationSettings().model_dump()
        await db.settings.insert_one(default)
        return default
    return s


@router.put("/settings/organization")
async def update_org_settings(body: dict,
                                 current: dict = Depends(require_roles(["admin"]))):
    body.pop("id", None); body["updated_at"] = now_iso()
    await db.settings.update_one({"id": "org_settings"}, {"$set": body}, upsert=True)
    await log_activity(current, "settings.updated", "Organization settings updated")
    s = await db.settings.find_one({"id": "org_settings"}, {"_id": 0})
    return s


# ================================================================
# SAVED FILTERS
# ================================================================
@router.get("/filters/saved")
async def list_saved_filters(entity: str = "candidates",
                                current: dict = Depends(get_current_user)):
    items = await db.saved_filters.find(
        {"user_id": current["id"], "entity": entity},
        {"_id": 0}).sort("created_at", -1).to_list(50)
    return items


@router.post("/filters/saved")
async def save_filter(name: str = Form(...), entity: str = Form("candidates"),
                        filters: str = Form(...),  # JSON string
                        current: dict = Depends(get_current_user)):
    import json
    parsed = json.loads(filters)
    f = SavedFilter(user_id=current["id"], name=name, entity=entity, filters=parsed)
    await db.saved_filters.insert_one(f.model_dump())
    return f.model_dump()


@router.delete("/filters/saved/{fid}")
async def delete_saved_filter(fid: str, current: dict = Depends(get_current_user)):
    await db.saved_filters.delete_one({"id": fid, "user_id": current["id"]})
    return {"success": True}


# ================================================================
# EXCEL IMPORT (.xlsx)
# ================================================================
def _row_to_dict(headers, row):
    return {h: (r if r is not None else "") for h, r in zip(headers, row)}


@router.post("/candidates/import/xlsx/preview")
async def preview_xlsx(file: UploadFile = File(...),
                         current: dict = Depends(require_roles(["admin"]))):
    """Read Excel & return preview + duplicate flags. Does NOT import."""
    data = await file.read()
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {"headers": [], "rows": [], "total": 0, "duplicates": 0}
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    preview = []
    dups = 0
    for r in rows[1:51]:  # first 50 rows
        d = _row_to_dict(headers, r)
        email = str(d.get("email", "")).strip().lower()
        phone = str(d.get("phone", "")).strip()
        dup = False
        if email or phone:
            q = {"is_deleted": {"$ne": True}, "$or": []}
            if email: q["$or"].append({"email": email})
            if phone: q["$or"].append({"phone": phone})
            if q["$or"] and await db.candidates.find_one(q):
                dup = True; dups += 1
        preview.append({**d, "_duplicate": dup})
    return {"headers": headers, "rows": preview,
              "total": max(0, len(rows) - 1), "duplicates": dups}


@router.post("/candidates/import/xlsx")
async def import_xlsx(file: UploadFile = File(...),
                       skip_duplicates: bool = Form(True),
                       current: dict = Depends(require_roles(["admin"]))):
    data = await file.read()
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {"imported": 0, "skipped": 0, "errors": []}
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    imported, skipped, errors = 0, 0, []
    for i, r in enumerate(rows[1:], start=2):
        d = _row_to_dict(headers, r)
        email = str(d.get("email", "")).strip().lower()
        name = str(d.get("full_name") or d.get("name") or "").strip()
        phone = str(d.get("phone", "")).strip()
        if not (email and name and phone):
            skipped += 1; continue
        if skip_duplicates:
            if await db.candidates.find_one({"is_deleted": {"$ne": True},
                "$or": [{"email": email}, {"phone": phone}]}):
                skipped += 1; continue
        try:
            skills = [s.strip() for s in str(d.get("skills", "")).split(",") if s.strip()]
            c = Candidate(
                full_name=name, email=email, phone=phone,
                city=str(d.get("city") or "") or None,
                state=str(d.get("state") or "") or None,
                skills=skills,
                total_experience_years=float(d.get("experience") or d.get("total_experience_years") or 0),
                expected_salary=float(d.get("expected_salary") or 0) or None,
                current_salary=float(d.get("current_salary") or 0) or None,
                reference_name=str(d.get("reference_name") or "") or None,
                reference_phone=str(d.get("reference_phone") or "") or None,
                status="pending_verification",
            )
            c.candidate_code = await next_code("candidates", "UGS-C")
            c.profile_completion = compute_profile_completion(c.model_dump())
            await db.candidates.insert_one(c.model_dump())
            imported += 1
        except Exception as e:
            errors.append({"row": i, "error": str(e)})
            skipped += 1
    await log_activity(current, "candidates.imported_xlsx",
                        f"Imported {imported} from Excel ({skipped} skipped)")
    return {"imported": imported, "skipped": skipped, "errors": errors[:20]}


# ================================================================
# EXCEL EXPORT (.xlsx)
# ================================================================
@router.get("/candidates/export/xlsx")
async def export_candidates_xlsx(
    status: str = "", employee_id: str = "", partner_id: str = "",
    payment_status: str = "",
    current: dict = Depends(require_roles(["admin", "employee"])),
):
    filt = {"is_deleted": {"$ne": True}}
    if current["role"] == "employee":
        filt["assigned_employee_id"] = current["id"]
    if status: filt["status"] = status
    if employee_id: filt["assigned_employee_id"] = employee_id
    if partner_id: filt["partner_id"] = partner_id
    if payment_status: filt["payment_status"] = payment_status
    rows = await db.candidates.find(filt, {"_id": 0}).to_list(20000)

    wb = Workbook(); ws = wb.active; ws.title = "Candidates"
    headers = ["Code", "Name", "Email", "Phone", "Status", "Payment",
                 "Amount Paid", "Fee", "Skills", "Experience (yrs)",
                 "Expected Salary", "City", "Recruiter", "Company", "Partner",
                 "Reference Name", "Reference Phone", "Created"]
    ws.append(headers)
    # id -> name maps
    users = {u["id"]: u["full_name"] async for u in db.users.find({}, {"id": 1, "full_name": 1})}
    cos = {c["id"]: c["name"] async for c in db.companies.find({}, {"id": 1, "name": 1})}
    parts = {p["id"]: p["name"] async for p in db.partners.find({}, {"id": 1, "name": 1})}

    for c in rows:
        ws.append([
            c.get("candidate_code"), c.get("full_name"), c.get("email"),
            c.get("phone"), c.get("status"), c.get("payment_status"),
            c.get("amount_paid", 0), c.get("registration_fee", 0),
            ", ".join(c.get("skills") or []),
            c.get("total_experience_years") or 0,
            c.get("expected_salary") or 0,
            c.get("city") or "",
            users.get(c.get("assigned_employee_id")) or "",
            cos.get(c.get("assigned_company_id")) or "",
            parts.get(c.get("partner_id")) or "",
            c.get("reference_name") or "",
            c.get("reference_phone") or "",
            c.get("created_at") or "",
        ])
    # auto column width
    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    await log_activity(current, "candidates.exported_xlsx",
                        f"Exported {len(rows)} candidates to Excel")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=candidates.xlsx"})


# ================================================================
# LOGIN / LOGOUT activity (called from frontend)
# ================================================================
@router.post("/auth/log-login")
async def log_login(current: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current["id"]})
    await log_activity(current if not user else {**user, **current},
                        "user.login",
                        f"{user.get('full_name') or current['email']} logged in")
    return {"ok": True}


@router.post("/auth/log-logout")
async def log_logout(current: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current["id"]})
    await log_activity(current if not user else {**user, **current},
                        "user.logout",
                        f"{user.get('full_name') or current['email']} logged out")
    return {"ok": True}


# ================================================================
# EXTENDED GLOBAL SEARCH (adds employees, batches, partners)
# ================================================================
@router.get("/search/full")
async def search_full(q: str, current: dict = Depends(get_current_user)):
    if not q or len(q) < 2:
        return {"candidates": [], "companies": [], "jobs": [],
                 "employees": [], "batches": [], "partners": []}
    r = {"$regex": q, "$options": "i"}
    cand = await db.candidates.find(
        {"is_deleted": {"$ne": True},
          "$or": [{"full_name": r}, {"email": r}, {"phone": r},
                    {"candidate_code": r}, {"skills": r}]},
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "phone": 1,
          "candidate_code": 1, "status": 1}).limit(10).to_list(10)
    cos = await db.companies.find({"name": r},
        {"_id": 0, "id": 1, "name": 1, "industry": 1}).limit(10).to_list(10)
    jobs = await db.jobs.find({"$or": [{"title": r}, {"skills": r}]},
        {"_id": 0, "id": 1, "title": 1, "company_id": 1, "status": 1}).limit(10).to_list(10)
    emps = await db.users.find({"role": "employee",
        "$or": [{"full_name": r}, {"email": r}]},
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "designation": 1}).limit(10).to_list(10)
    batches = await db.batches.find({"$or": [{"name": r}, {"technology": r}]},
        {"_id": 0, "id": 1, "name": 1, "technology": 1, "status": 1}).limit(10).to_list(10)
    partners = await db.partners.find({"$or": [{"name": r}, {"phone": r},
        {"partner_code": r}]},
        {"_id": 0, "id": 1, "name": 1, "phone": 1, "partner_code": 1}).limit(10).to_list(10)
    return {"candidates": cand, "companies": cos, "jobs": jobs,
             "employees": emps, "batches": batches, "partners": partners}


# ================================================================
# NOTE: Main list endpoint /api/candidates (in server.py) already supports
# partner_id + include_archived + reference_name search.
# ================================================================
